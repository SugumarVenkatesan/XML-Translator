from flask import request, Flask
from flask.views import MethodView
from flask import render_template
from xml.etree import ElementTree as ET
                                     
import sys
import os
import json
import openpyxl
import re

app = Flask("XML Translator")
base_dir = os.getcwd()
lookup_data = os.path.join(base_dir, "config")

def read_lookup_excel(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    data = []
    sheet = wb.active
    for i in range(2,sheet.max_row+1):
        english = sheet.cell(row=i,column=1).value
        japanese = sheet.cell(row=i,column=2).value
        data.append((english,japanese))
    return data

class XMLTranlsatorHome(MethodView):
 
    def get(self):
        data = read_lookup_excel(lookup_data+os.sep+'EN-JP-lookup.xlsx')
        return render_template("base.html",lookup_data = data)
    
    
    def post(self):
        try:
            source_xml = request.form["source_textarea"]
            if not source_xml:
                raise Exception("Please provide the source xml to translate")
            try:
                ET.fromstring(source_xml)
            except:
                raise Exception("Invalid xml")
            
            tags_trans = request.form.getlist("tags_trans")
            
            if not tags_trans:
                raise Exception("Please select the xml tags to translate")
                 
            source_trans = request.form["source_trans"]
            dest_trans = request.form["dest_trans"]
            
            if source_trans == dest_trans:
                raise Exception("Translation Language should be different")
            
            source_index = 0 if source_trans == "english" else 1
            dest_index = 0 if source_trans == "japanese" else 1
            
            dest_xml = source_xml
            
            cloned_values = []
            
            data = read_lookup_excel(lookup_data+os.sep+'EN-JP-lookup.xlsx')
            for each_tag in tags_trans:
                pat = "<%s>(.*?)</%s>" % (each_tag,each_tag)
                rl_pat = "<%s>%s</%s>"
                values = re.findall(pat,dest_xml)
                cloned_values = values.copy()
                for each_value in values:
                    for each_dt in data:
                        if each_value == each_dt[source_index]:
                            dest_xml = dest_xml.replace(rl_pat % (each_tag,each_value,each_tag),
                                               rl_pat % (each_tag,each_value.replace(each_dt[source_index],each_dt[dest_index]),each_tag))
                            cloned_values.remove(each_value)
            return json.dumps({"dest_trans":dest_xml,
                               "success":"The Translation is successfully completed",
                               "warning":"Could not find translation lookup for %s" %(",".join(cloned_values)) if cloned_values else "",
                               "error":""})
        except:
            return json.dumps({"dest_trans":"",
                               "success":"",
                               "warning":"",
                               "error":sys.exc_info()[1].__str__()})
        
        
        
        
        
         
         
 
app.add_url_rule("/", view_func=XMLTranlsatorHome.as_view("XMLTranlsatorHome"))
app.run(host="0.0.0.0", debug=True)